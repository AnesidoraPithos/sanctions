"""
Export Routes

POST /api/export/{search_id}  — export results as JSON, Excel, or PDF.
"""

import io
import logging
import os
import sys
from typing import Any, Dict

from fastapi import APIRouter, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

# Ensure backend modules are importable
backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

from db_operations.db import get_search_results

logger = logging.getLogger(__name__)
router = APIRouter()


class ExportRequest(BaseModel):
    format: str  # "json" | "excel" | "pdf"


# ---------------------------------------------------------------------------
# Helper: build Excel workbook
# ---------------------------------------------------------------------------

def _build_excel(results: Dict[str, Any]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is not installed") from exc

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a2744")

    summary_rows = [
        ("Field", "Value"),
        ("Entity Name", results.get("entity_name", "")),
        ("Tier", results.get("tier", "")),
        ("Risk Level", results.get("risk_level", "")),
        ("Sanctions Hits", results.get("sanctions_hits", 0)),
        ("Timestamp", str(results.get("timestamp", ""))),
    ]
    for r_idx, row in enumerate(summary_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40

    # --- Sheet 2: Sanctions ---
    ws2 = wb.create_sheet("Sanctions")
    sanctions_headers = ["Name", "List", "Type", "Address", "Match Quality", "Score"]
    for c_idx, h in enumerate(sanctions_headers, start=1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, hit in enumerate(results.get("sanctions_data", []), start=2):
        ws2.cell(row=r_idx, column=1, value=hit.get("name", ""))
        ws2.cell(row=r_idx, column=2, value=hit.get("list", ""))
        ws2.cell(row=r_idx, column=3, value=hit.get("type", ""))
        ws2.cell(row=r_idx, column=4, value=hit.get("address", ""))
        ws2.cell(row=r_idx, column=5, value=hit.get("match_quality", ""))
        ws2.cell(row=r_idx, column=6, value=hit.get("combined_score", ""))

    # --- Sheet 3: Subsidiaries ---
    fi = results.get("financial_intelligence") or {}
    subs = results.get("subsidiaries", [])
    ws3 = wb.create_sheet("Subsidiaries")
    sub_headers = ["Name", "Jurisdiction", "Status", "Level", "Sanctions Hits"]
    for c_idx, h in enumerate(sub_headers, start=1):
        cell = ws3.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, sub in enumerate(subs, start=2):
        ws3.cell(row=r_idx, column=1, value=sub.get("name", ""))
        ws3.cell(row=r_idx, column=2, value=sub.get("jurisdiction", ""))
        ws3.cell(row=r_idx, column=3, value=sub.get("status", ""))
        ws3.cell(row=r_idx, column=4, value=sub.get("level", ""))
        ws3.cell(row=r_idx, column=5, value=sub.get("sanctions_hits", 0))

    # --- Sheet 4: Directors ---
    ws4 = wb.create_sheet("Directors")
    dir_headers = ["Name", "Title", "Nationality", "Sanctions Hits"]
    for c_idx, h in enumerate(dir_headers, start=1):
        cell = ws4.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, d in enumerate(fi.get("directors", []), start=2):
        ws4.cell(row=r_idx, column=1, value=d.get("name", ""))
        ws4.cell(row=r_idx, column=2, value=d.get("title", ""))
        ws4.cell(row=r_idx, column=3, value=d.get("nationality", ""))
        ws4.cell(row=r_idx, column=4, value=d.get("sanctions_hits", 0))

    # --- Sheet 5: Shareholders ---
    ws5 = wb.create_sheet("Shareholders")
    sh_headers = ["Name", "Type", "Ownership %", "Jurisdiction", "Sanctions Hits"]
    for c_idx, h in enumerate(sh_headers, start=1):
        cell = ws5.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, sh in enumerate(fi.get("shareholders", []), start=2):
        ws5.cell(row=r_idx, column=1, value=sh.get("name", ""))
        ws5.cell(row=r_idx, column=2, value=sh.get("type", ""))
        ws5.cell(row=r_idx, column=3, value=sh.get("ownership_percentage", ""))
        ws5.cell(row=r_idx, column=4, value=sh.get("jurisdiction", ""))
        ws5.cell(row=r_idx, column=5, value=sh.get("sanctions_hits", 0))

    # --- Sheet 6: Transactions ---
    ws6 = wb.create_sheet("Transactions")
    tx_headers = ["Type", "Counterparty", "Amount", "Currency", "Date", "Purpose"]
    for c_idx, h in enumerate(tx_headers, start=1):
        cell = ws6.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, tx in enumerate(fi.get("transactions", []), start=2):
        ws6.cell(row=r_idx, column=1, value=tx.get("transaction_type", ""))
        ws6.cell(row=r_idx, column=2, value=tx.get("counterparty", ""))
        ws6.cell(row=r_idx, column=3, value=tx.get("amount", ""))
        ws6.cell(row=r_idx, column=4, value=tx.get("currency", ""))
        ws6.cell(row=r_idx, column=5, value=tx.get("transaction_date", ""))
        ws6.cell(row=r_idx, column=6, value=tx.get("purpose", ""))

    # --- Sheet 7: Management Network (Phase 4) ---
    network_data = results.get("network_data") or {}
    director_pivots = network_data.get("director_pivots") or results.get("director_pivots") or []
    ws7 = wb.create_sheet("Management Network")
    mn_headers = ["Director", "Title", "Interlocked Company", "Filing Date", "Source URL"]
    for c_idx, h in enumerate(mn_headers, start=1):
        cell = ws7.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    r_idx = 2
    for pivot in director_pivots:
        director_name = pivot.get("director_name", "") if isinstance(pivot, dict) else ""
        title = pivot.get("title", "") if isinstance(pivot, dict) else ""
        companies = pivot.get("companies", []) if isinstance(pivot, dict) else []
        for company in companies:
            ws7.cell(row=r_idx, column=1, value=director_name)
            ws7.cell(row=r_idx, column=2, value=title)
            ws7.cell(row=r_idx, column=3, value=company.get("company_name", "") if isinstance(company, dict) else "")
            ws7.cell(row=r_idx, column=4, value=company.get("filing_date", "") if isinstance(company, dict) else "")
            ws7.cell(row=r_idx, column=5, value=company.get("source_url", "") if isinstance(company, dict) else "")
            r_idx += 1

    # --- Sheet 8: Infrastructure (Phase 4) ---
    infrastructure = network_data.get("infrastructure") or results.get("infrastructure") or []
    ws8 = wb.create_sheet("Infrastructure")
    infra_headers = ["Domain", "Registrant Org", "Registrar", "Created", "Nameservers", "Related Entities"]
    for c_idx, h in enumerate(infra_headers, start=1):
        cell = ws8.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, item in enumerate(infrastructure, start=2):
        if not isinstance(item, dict):
            continue
        ws8.cell(row=r_idx, column=1, value=item.get("domain", ""))
        ws8.cell(row=r_idx, column=2, value=item.get("registrant_org", ""))
        ws8.cell(row=r_idx, column=3, value=item.get("registrar", ""))
        ws8.cell(row=r_idx, column=4, value=item.get("creation_date", ""))
        nameservers = item.get("nameservers") or []
        ws8.cell(row=r_idx, column=5, value=", ".join(nameservers) if isinstance(nameservers, list) else "")
        related = item.get("related_entities") or []
        ws8.cell(row=r_idx, column=6, value=", ".join(related) if isinstance(related, list) else "")

    # --- Sheet 9: Beneficial Owners (Phase 4) ---
    beneficial_owners = network_data.get("beneficial_owners") or results.get("beneficial_owners") or []
    ws9 = wb.create_sheet("Beneficial Owners")
    bo_headers = ["Name", "Nationality", "Ownership %", "Source", "Verification Date"]
    for c_idx, h in enumerate(bo_headers, start=1):
        cell = ws9.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r_idx, owner in enumerate(beneficial_owners, start=2):
        if not isinstance(owner, dict):
            continue
        ws9.cell(row=r_idx, column=1, value=owner.get("name", ""))
        ws9.cell(row=r_idx, column=2, value=owner.get("nationality", ""))
        ws9.cell(row=r_idx, column=3, value=owner.get("ownership_pct", ""))
        ws9.cell(row=r_idx, column=4, value=owner.get("source", ""))
        ws9.cell(row=r_idx, column=5, value=owner.get("verification_date", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Helper: build PDF report
# ---------------------------------------------------------------------------

def _build_pdf(results: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="reportlab is not installed") from exc

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], textColor=colors.HexColor("#1a4fbd"))
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], textColor=colors.HexColor("#1a4fbd"))
    body_style = styles["BodyText"]

    elements = []

    # Cover
    elements.append(Paragraph("Entity Background Research Report", title_style))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a4fbd")))
    elements.append(Spacer(1, 0.5 * cm))

    entity_name = results.get("entity_name", "Unknown")
    risk_level = results.get("risk_level", "UNKNOWN")
    tier = results.get("tier", "base")
    timestamp = str(results.get("timestamp", ""))

    summary_data = [
        ["Entity Name", entity_name],
        ["Tier", tier.upper()],
        ["Risk Level", risk_level],
        ["Sanctions Hits", str(results.get("sanctions_hits", 0))],
        ["Generated", timestamp],
    ]
    tbl = Table(summary_data, colWidths=[4 * cm, 13 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8edf7")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 0.8 * cm))

    # Sanctions table
    sanctions = results.get("sanctions_data", [])
    if sanctions:
        elements.append(Paragraph("Sanctions Hits", h2_style))
        elements.append(Spacer(1, 0.2 * cm))
        s_data = [["Name", "List", "Match Quality", "Score"]]
        for hit in sanctions[:50]:
            s_data.append([
                hit.get("name", "")[:50],
                hit.get("list", "")[:30],
                hit.get("match_quality", ""),
                f"{hit.get('combined_score', 0):.0f}%",
            ])
        s_tbl = Table(s_data, colWidths=[6 * cm, 5 * cm, 3 * cm, 3 * cm])
        s_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a4fbd")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fb")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(s_tbl)
        elements.append(Spacer(1, 0.8 * cm))

    # Intelligence report
    report_text = results.get("intelligence_report")
    if report_text:
        elements.append(Paragraph("Intelligence Report", h2_style))
        elements.append(Spacer(1, 0.2 * cm))
        # Strip markdown-ish formatting
        clean = report_text.replace("**", "").replace("##", "").replace("# ", "")
        for para in clean.split("\n\n"):
            para = para.strip()
            if para:
                elements.append(Paragraph(para[:2000], body_style))
                elements.append(Spacer(1, 0.2 * cm))

    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Route
# ---------------------------------------------------------------------------

@router.post("/{search_id}")
async def export_results(search_id: str, body: ExportRequest):
    """Export search results as JSON, Excel, or PDF."""
    fmt = body.format.lower()
    if fmt not in ("json", "excel", "pdf"):
        raise HTTPException(status_code=400, detail="format must be 'json', 'excel', or 'pdf'")

    results = get_search_results(search_id)
    if not results:
        raise HTTPException(status_code=404, detail=f"Search {search_id} not found")

    entity_slug = results.get("entity_name", search_id).replace(" ", "_")[:40]

    if fmt == "json":
        return JSONResponse(
            content=results,
            headers={"Content-Disposition": f'attachment; filename="{entity_slug}_{search_id[:8]}.json"'},
        )

    if fmt == "excel":
        data = _build_excel(results)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{entity_slug}_{search_id[:8]}.xlsx"'},
        )

    # pdf
    data = _build_pdf(results)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{entity_slug}_{search_id[:8]}.pdf"'},
    )
